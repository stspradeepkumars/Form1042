import openpyxl
import  pyodbc
import  os
import  datetime 
from datetime import datetime as dt
import  glob
import time as t
import shutil
from faker import Faker

def delh():
    os.remove("MyFile.txt")

    
def delimg():
    #pattern = "selenium-screenshot*"
    pattern = "*.png"
    files = glob.glob(pattern)
    for file in files:
        os.remove(file)
  

def delht():
    os.remove("My.html")  

def writeh(qry):
    qry1= qry.replace("'","")
    val=qry1.split(",")
    file1 = open("MyFile.txt","a")
    file1.write('<tr><th style="text-align: left">'+ val[1] +'</th><th style="text-align: left">'+val[2]+'</th><th>'+val[5]+'</th></tr>')
    file1.close()

def projectDirectory():
    path = os.getcwd()
    folder = path.split("\\")
    projectDirectory=""
    for i in range(0,len(folder)):
        projectDirectory=projectDirectory+folder[i]+"\\"
        if  folder[i]=='94xrevamp':
            break
    return  projectDirectory  

def currentDateAndTime():
    currentDateAndTime = dt.now()
    currentDateAndTime = str(currentDateAndTime)
    currentDateAndTime = currentDateAndTime[8:10]+currentDateAndTime[5:7]+currentDateAndTime[2:4]+currentDateAndTime[11:13]+currentDateAndTime[14:16]+currentDateAndTime[17:19]
    return  currentDateAndTime
    
class associatelib:

 def fndbget(self,qry):
    try:
        connect=pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};server=143.47.101.114;Database=QA_Automation_Staging_DB;uid=StagingAutomationDB_User;pwd=Asc$nd@s#113;Trusted_Connection=no;')
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(sqlstate)
    else:
        cursor=connect.cursor()
        cursor.execute(qry)
        for row in cursor:
            return (row[0])
        cursor.close()
        connect.close()

 def persget(self,qry):
    connect=pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};server=143.47.101.114;Database=QA_Automation_Staging_DB;uid=StagingAutomationDB_User;pwd=Asc$nd@s#113;Trusted_Connection=no;')
    cursor=connect.cursor()
    ret=cursor.execute(qry)
    records=ret.fetchall()
    for row in records:
      return(records)
    cursor.close()
    connect.close()



 def writelog(self,qry):
    connect=pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};server=143.47.101.114;Database=QA_Automation_Staging_DB;uid=StagingAutomationDB_User;pwd=Asc$nd@s#113;Trusted_Connection=no;')
    cursor=connect.cursor()
    cursor.execute("Insert into tb_log (logid,tsid,tsdesc,expectedresult,actualresult,runstatus,reference) values ("+qry+");")
    connect.commit()
    cursor.close()
    connect.close()
    writeh(qry)

 def logid(self):
    delimg()
    try:
        delh()
    except:
        pass
    return("Log_"+"{:%Y%m%d%H%M%S}".format(datetime.datetime.now()))
    
    #all_files = os.listdir()
    #for f in all_files:
    #    print(f)
        #os.remove(f)
    #for f in glob.glob("*.png"):
    #    os.remove(f)

 def createh(self):
    starttime=os.environ['starttime']
    browser=os.environ['browser']
    EmailAddress=os.environ['EmailAddress']
    logid=os.environ['logid']
    Password=os.environ['Password']
    url=os.environ['url']
    end_time = t.localtime()
    end_time = t.strftime("%H:%M:%S", end_time)
    end_time= str(end_time)
    t1 = dt.strptime(starttime, "%H:%M:%S")
    t2 = dt.strptime(end_time, "%H:%M:%S")
    delta  = t2 - t1
    sec = delta.total_seconds()
    min = sec / 60
    duration = min.__round__(2)
    duration = str(duration)
    strstatus = 'Fail'
    file3 = open("MyFile.txt","r")
    valread=file3.read()
    passc = valread.count("<th>Pass</th></tr>")
    failc = valread.count("<th>Fail</th></tr>")
    totalc = valread.count('<tr>')
    if failc==0 :
        strstatus = 'Pass'
    sumtable='<table cellspacing=”2” width=”100%” border=”1”><tr><th style="background-color:#FFA500">Log ID</th><th style="background-color:#FFA500" >Run Status</th><th style="background-color:#FFA500">Total Testcases</th><th style="background-color:#FFA500">Pass TC</th><th style="background-color:#FFA500">Fail TC</th><th style="background-color:#FFA500">Start Time</th><th style="background-color:#FFA500">End Time</th><th style="background-color:#FFA500">Duration</th></tr><tr><th>'+logid+'</th><th>'+strstatus+'</th><th>'+str(totalc)+'</th><th>'+str(passc)+'</th><th>'+str(failc)+'</th><th>'+str(starttime)+'</th><th>'+str(end_time)+'</th><th>'+str(duration)+'</th></tr></table>'
    #sumtable='<table cellspacing=”2” width=”100%” border=”1”><tr><th style="background-color:#FFA500">Log ID</th><th style="background-color:#FFA500" >Browser</th><th style="background-color:#FFA500" >Run Status</th><th style="background-color:#FFA500">Total Testcases</th><th style="background-color:#FFA500">Pass TC</th><th style="background-color:#FFA500">Fail TC</th></tr><tr><th>'+logid+'</th><th>'+@{Browser}+'</th><th>'+strstatus+'</th><th>'+str(totalc)+'</th><th>'+str(passc)+'</th><th>'+str(failc)+'</th></tr></table>'
    valuestr='<html> <style> table, th, td {   border:1px solid black; } </style> <body>  <h1>Run Report</h1>  <h2>Run Summary</h2> '+ sumtable +'<h2>Run Details</h2><p><b>Url and region: </b>'+url+' / STAGING</p><p><b>Credentials Used: </b>'+EmailAddress+' / '+Password+'</p> <p><b>Browser Used: </b>'+browser+'</p> <table style="width:50%" > <tr><th style="background-color:#FA58F4" >Test Scenario ID</th><th style="background-color:#FA58F4">Test Scenario Description</th><th style="background-color:#FA58F4">Test Status</th></tr>'+ valread  +'</table> </body> </html>'
    valuestr = valuestr.replace(">Pass", ' style="color: green">Pass')
    valuestr = valuestr.replace(">'Pass'", ' style="color: green">Pass')
    valuestr = valuestr.replace(">Fail", ' style="color: red">Fail')
    valuestr = valuestr.replace(">'Fail'", ' style="color: red">Fail')
    file3.close()
    file2 = open("My.html","w")
    file2.write(valuestr)
    file2.close()
    #delh()
  
#  def createh(self,logid):
#     strstatus = 'Fail'
#     file3 = open("MyFile.txt","r")
#     valread=file3.read()
#     passc = valread.count("<th>Pass</th></tr>")
#     failc = valread.count("<th>Fail</th></tr>")
#     totalc = valread.count('<tr>')
#     if failc==0 :
#         strstatus = 'Pass'
#     sumtable='<table cellspacing=”2” width=”100%” border=”1”><tr><th style="background-color:#FFA500">Log ID</th><th style="background-color:#FFA500" >Run Status</th><th style="background-color:#FFA500">Total Testcases</th><th style="background-color:#FFA500">Pass TC</th><th style="background-color:#FFA500">Fail TC</th></tr><tr><th>'+logid+'</th><th>'+strstatus+'</th><th>'+str(totalc)+'</th><th>'+str(passc)+'</th><th>'+str(failc)+'</th></tr></table>'
#     #sumtable='<table cellspacing=”2” width=”100%” border=”1”><tr><th style="background-color:#FFA500">Log ID</th><th style="background-color:#FFA500" >Browser</th><th style="background-color:#FFA500" >Run Status</th><th style="background-color:#FFA500">Total Testcases</th><th style="background-color:#FFA500">Pass TC</th><th style="background-color:#FFA500">Fail TC</th></tr><tr><th>'+logid+'</th><th>'+@{Browser}+'</th><th>'+strstatus+'</th><th>'+str(totalc)+'</th><th>'+str(passc)+'</th><th>'+str(failc)+'</th></tr></table>'
#     valuestr='<html> <style> table, th, td {   border:1px solid black; } </style> <body>  <h1>Run Report</h1>  <h2>Run Summary</h2> '+ sumtable +'<h2>Run Details</h2> <table style="width:50%" > <tr><th style="background-color:#FA58F4" >Test Scenario ID</th><th style="background-color:#FA58F4">Test Scenario Description</th><th style="background-color:#FA58F4">Test Status</th></tr>'+ valread  +'</table> </body> </html>'
#     valuestr = valuestr.replace(">Pass", ' style="color: green">Pass')
#     valuestr = valuestr.replace(">'Pass'", ' style="color: green">Pass')
#     valuestr = valuestr.replace(">Fail", ' style="color: red">Fail')
#     valuestr = valuestr.replace(">'Fail'", ' style="color: red">Fail')
#     file3.close()
#     file2 = open("My.html","w")
#     file2.write(valuestr)
#     file2.close()
#     delh()

 def readExcel(self,Sheet,ColumnHeader,Row):
    Row = int(Row)
    path = "../../Test Data/Form941.xlsx"
    book = openpyxl.load_workbook(path)
    worksheet = book[Sheet]
    total_column = worksheet.max_column
    cellValue = None
    for i in range(1,total_column+1):  
        x3 = worksheet.cell(row=1, column=i).value
        if x3==ColumnHeader:
            cellValue = i
    value = worksheet.cell(row=Row, column=cellValue).value
    return value

 def time(self):
    end_time = t.localtime()
    end_time = t.strftime("%H:%M:%S", end_time)
    return end_time
 
 def removePabotResult(self):
    try:
        directory = "pabot_results"
        parent = "D:\\94x_Multiple Scenario\\Test suite\\941SMOKE\\"
        path = os.path.join(parent, directory)
        os.rmdir(path)
    except:
       pass
    
 def downloadedFile(self,oldFiles,newFiles):
    for i in newFiles[:]:
        if i in oldFiles:
            newFiles.remove(i)
    return  newFiles[0]
 
 def filelist(self):
    downloads_path = os.path.expanduser("C:\\Users\\STS717-PUVANASRI J\\Downloads")
    files= os.listdir(downloads_path)
    return  files
 def timesub(self, starttime, endtime):
    t1 = dt.strptime(starttime, "%H:%M:%S")
    t2 = dt.strptime(endtime, "%H:%M:%S")
    delta  = t2 - t1
    sec = delta.total_seconds()
    return sec

 def BackupFiles(self, location):
     current_date_time=currentDateAndTime()
     backupFolder = location+"//Results//Backup//"+current_date_time
     os.makedirs(backupFolder)
     if  os.path.exists(location+"//My.html"):
            shutil.move(location+"//My.html",backupFolder)
     if  os.path.exists(location+"//log.html"):
             shutil.move(location+"//log.html",backupFolder)
     if  os.path.exists(location+"//support//reference"):
             shutil.move(location+"//support//reference",backupFolder)
     return  location


 def location_path(self, source_path, destination_path):
    path = os.path.realpath(__file__)
    dir = os.path.dirname(path)
    dir = dir.replace(source_path, destination_path)
    return dir